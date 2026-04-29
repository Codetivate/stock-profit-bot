"""
parse_set_zip.py — SET financial statement parser

Extracts net profit data from SET news zip files.
Format: XLSX with sheets like 'PL 10-11', 'BS 7-9', etc.

Usage:
    from parse_set_zip import parse_zip
    data = parse_zip("0737FIN250220261406350902T.zip")
"""
import os
import re
import zipfile
import tempfile
from typing import Optional, Dict, Any
from datetime import datetime
from dataclasses import dataclass, asdict

import openpyxl
import xlrd  # type: ignore  # legacy .xls reader


@dataclass
class FinancialData:
    """Extracted financial data from one report."""
    symbol: str
    filename: str
    period_label: str  # e.g. "งบประจำปี 2568"
    period_type: str   # "annual" | "quarterly" | "half"
    year: int          # Thai year, e.g. 2568
    quarter: Optional[int]  # 1, 2, 3, 4 or None for annual

    # Income statement numbers (in millions of baht)
    revenue: Optional[float] = None          # รวมรายได้
    revenue_prior: Optional[float] = None    # previous period
    net_profit: Optional[float] = None       # กำไรสำหรับปี/ไตรมาส
    net_profit_prior: Optional[float] = None
    shareholder_profit: Optional[float] = None  # ส่วนที่เป็นของผู้ถือหุ้น (3-month standalone)
    shareholder_profit_prior: Optional[float] = None
    # Cumulative-period profit from the "longer" PL sheet when present
    # (e.g. 9-month cumulative in a 9M filing, 6-month in an H1 filing,
    # full year in the FY filing). Q2 and Q4 for issuers that don't file
    # H1 are back-computed from these in compute_standalone_quarters.
    shareholder_profit_cum: Optional[float] = None
    shareholder_profit_cum_prior: Optional[float] = None
    cum_months: Optional[int] = None   # 3 / 6 / 9 / 12
    # Period length (in months) of the primary PL sheet — i.e. the
    # sheet that ``shareholder_profit`` was extracted from. 3 means we
    # have a real 3-month standalone; 6/9 means the filing only ships
    # the cumulative sheet, and ``shareholder_profit`` actually
    # equals ``shareholder_profit_cum`` rather than the standalone
    # quarter. compute_standalone_quarters uses this to avoid treating
    # a cum-only number as Q2/Q3 standalone.
    primary_months: Optional[int] = None
    eps: Optional[float] = None              # กำไรต่อหุ้น
    eps_prior: Optional[float] = None


def _is_revenue_row(text: str) -> bool:
    """Match 'รวมรายได้' (total revenue)."""
    if not text:
        return False
    text = str(text).strip()
    return text == "รวมรายได้"


def _is_netprofit_row(text: str) -> bool:
    """Match the income-statement bottom line in any of the variants
    SET filers actually use:

      - กำไรสำหรับ(ปี|งวด|ไตรมาส|รอบระยะเวลา)             ← profit case
      - กำไร (ขาดทุน) สำหรับ(...)                          ← spaced variant
      - กำไร(ขาดทุน)สำหรับ(...)                            ← no-space (AJ)
      - (ขาดทุน) กำไรสำหรับ(...)                           ← loss-leading (TMT loss qtrs)
      - (ขาดทุน)กำไรสำหรับ(...)                            ← loss-leading no-space
      - กำไรสุทธิสำหรับ... / กำไรสุทธิ / กำไร (ขาดทุน) สุทธิ ← banks

    Excludes the comprehensive-income row 'กำไรเบ็ดเสร็จรวม...' and
    its loss-leading variants — those would include OCI items, not
    just net profit. The minority-interest row is filtered separately
    by `_is_shareholder_profit_row`.
    """
    if not text:
        return False
    text = str(text).strip()
    # Reject discontinued-operations breakdowns. ITC 2564 splits
    # ``กำไรสำหรับปี`` into:
    #   r32: กำไรสำหรับปีจากการดำเนินงานต่อเนื่อง   (continuing only)
    #   r33: กำไรจากการดำเนินงานที่ยกเลิก          (discontinued only)
    #   r35: กำไรสำหรับปี                          (TOTAL — what SET uses)
    # The original regex matched r32 first because "^กำไรสำหรับปี"
    # passes regardless of the trailing qualifier, picking up only the
    # continuing-ops portion. We want the unqualified total — reject
    # any row that names a specific sub-population of operations.
    if (
        "จากการดำเนินงานต่อเนื่อง" in text
        or "จากการดำเนินงานที่ยกเลิก" in text
    ):
        return False
    # Inline-share rows like "กำไรสำหรับปีส่วนที่เป็นของผู้ถือหุ้น
    # บริษัทฯ" or INTUCH 2565's "กำไรสำหรับปีส่วนที่เป็นของบริษัทใหญ่"
    # collapse the bottom line and the parent-share allocation into a
    # single label. They're already handled by _is_shareholder_profit_row
    # (the inline_share branch). Reject them here so the elif chain in
    # the main loop assigns them to result.shareholder_profit instead
    # of result.net_profit — otherwise the parent row gets consumed by
    # net_profit and shareholder falls through to the cross-sheet
    # recovery loop, which can latch onto a year-header value in the
    # statement-of-equity sheet (FSX 2568 was reading sh = 0.002568,
    # i.e. "year 2568" / 1,000,000).
    if "ส่วนที่เป็น" in text or "ส่วนของ" in text:
        return False
    # The label group accepts both Thai phrasings for the period unit
    # ("ปี/ไตรมาส/งวด" and the longer "รอบระยะเวลา").
    period_grp = r"(ปี|งวด|ไตรมาส|รอบระยะเวลา)"
    # Optional "(ขาดทุน)" with optional surrounding whitespace, in
    # either prefix or infix position. \s* lets us tolerate the same
    # phrase with or without spaces around the parens AND inside them —
    # BEAUTY ships rows like ``กำไร (ขาดทุน ) สำหรับงวด`` with an
    # extra space before the closing paren, which broke the original
    # `\(ขาดทุน\)` (zero-space-only) match and silently dropped the
    # net-profit row for every BEAUTY filing.
    loss_prefix = r"\(\s*ขาดทุน\s*\)\s*"      # "(ขาดทุน) " (TMT loss qtr)
    loss_infix = r"\s*\(\s*ขาดทุน\s*\)\s*"    # " (ขาดทุน) " (CPALL et al.)
    patterns = [
        rf"^กำไรสำหรับ{period_grp}",
        rf"^กำไร{loss_infix}สำหรับ{period_grp}",
        rf"^{loss_prefix}กำไรสำหรับ{period_grp}",
        # Bare "ขาดทุนสำหรับ..." with no preceding กำไร — used by issuers
        # who book a full-year loss (AJ 2567, 2568) rather than the
        # parenthesised "(ขาดทุน) กำไร..." form. Specifically targets the
        # period unit afterwards so it doesn't catch operating-loss rows
        # like "ขาดทุนจากกิจกรรมดำเนินงาน".
        rf"^ขาดทุนสำหรับ{period_grp}",
        r"^กำไรสุทธิสำหรับ",
        r"^ขาดทุนสุทธิสำหรับ",
        r"^กำไรสุทธิ$",
        r"^ขาดทุนสุทธิ$",
        r"^กำไร\s*\(ขาดทุน\)\s*สุทธิ$",
        rf"^{loss_prefix}กำไรสุทธิ$",
    ]
    return any(re.match(p, text) for p in patterns)


def _is_shareholder_profit_row(text: str) -> bool:
    """Match the row that splits net profit into the parent-company
    share vs minority interest.

    Covers all variations seen in SET XLSX filings:
      - ส่วนที่เป็นของผู้ถือหุ้นของบริษัท          (CPALL, most non-financial)
      - ส่วนที่เป็นของบริษัทใหญ่                    (SCB, KKP)
      - ส่วนที่เป็นของธนาคาร                         (KBANK, BBL, KTB, TTB, BAY)
      - ส่วนที่เป็นของผู้ถือหุ้นของธนาคาร          (CIMBT)
      - ส่วนของผู้เป็นเจ้าของของบริษัทใหญ่         (TMT, used in many newer filings)
      - ส่วนของบริษัทใหญ่                           (THE, some others)

    Explicitly rejects the sibling row for minority interest
    (ส่วนที่เป็นของส่วนได้เสียที่ไม่มีอำนาจควบคุม or
    ส่วนของส่วนได้เสียที่ไม่มีอำนาจควบคุม) so we never
    accidentally capture the wrong number. The minority row is
    distinguished by ``ส่วนได้เสีย`` or ``ไม่มีอำนาจ`` markers.

    SET's company-highlights API reports this parent-share figure
    (e.g. TMT 2566: 333.88 MB) — not the consolidated total
    (332.74 MB). Matching this row precisely is what keeps our
    quarterly history aligned with what users see on SET's site.
    """
    if not text:
        return False
    text = str(text).strip()
    # Strip leading bullet markers — ILINK 2564 FY indents parent-share
    # rows with a literal hyphen prefix (``- ส่วนที่เป็นของ…``) which
    # otherwise blocks the ``startswith("ส่วน…")`` check.
    text = text.lstrip("-*•◦· ").strip()
    # Both "ส่วนที่เป็น..." and "ส่วนของ..." are used in the wild.
    # INTUCH 2565 FY also collapses the line into one phrase
    # (``กำไรสำหรับปีส่วนที่เป็นของบริษัทใหญ่``) — match those too,
    # but only when the row clearly references the period total
    # (``กำไรสำหรับ(ปี|งวด)``) so we don't catch random labels that
    # happen to contain ``ส่วนที่เป็น``.
    # WAVE 2567 FY splits the parent-share line by continuing /
    # discontinued operations and emits the section sum row as
    # ``รวมส่วนที่เป็นของบริษัทใหญ่``. We need to keep this row — it's
    # the parent total — while still rejecting the balance-sheet
    # equity total ``รวมส่วนของบริษัทใหญ่`` (no ``ที่เป็น``) handled
    # below. The "starts with" check therefore covers both forms.
    starts_share = (
        text.startswith("ส่วนที่เป็น")
        or text.startswith("ส่วนของ")
        or text.startswith("รวมส่วนที่เป็น")
    )
    inline_share = (
        re.match(r"^กำไร(\s*\(\s*ขาดทุน\s*\)\s*)?สำหรับ(ปี|งวด|ไตรมาส|รอบระยะเวลา)",
                 text)
        and ("ส่วนที่เป็น" in text or "ส่วนของ" in text)
    )
    if not (starts_share or inline_share):
        return False
    # Minority-interest row has a distinct signature.
    if "ส่วนได้เสีย" in text or "ไม่มีอำนาจ" in text:
        return False
    # Balance-sheet equity rows look almost identical ("ส่วนของผู้ถือ
    # หุ้นของบริษัทฯ" — total parent shareholder EQUITY) and the
    # ``ส่วนของ`` prefix matches them too. PSL 2564 BS row 101 has this
    # exact label with a 14B baht equity total which would silently
    # masquerade as a 14,364 MB net profit. Real PL "profit allocation"
    # rows always say "ที่เป็นของ..." or "ของบริษัทใหญ่" / "ของผู้เป็น
    # เจ้าของ..." — never "ของผู้ถือหุ้น" without "ที่เป็น" in front.
    if text.startswith("ส่วนของผู้ถือหุ้น"):
        return False
    # Aggregate equity rows ("รวมส่วนของบริษัทใหญ่") are BS only.
    # The PL aggregate "รวมส่วนที่เป็นของบริษัทใหญ่" was already
    # admitted by ``starts_share`` above, so we only need to filter
    # the BS-equity form here (no ``ที่เป็น``).
    if text.startswith("รวมส่วนของ"):
        return False
    # Must reference the parent entity (company or bank).
    return "บริษัท" in text or "ธนาคาร" in text


def _is_eps_row(text: str) -> bool:
    """Match 'กำไรต่อหุ้น' (EPS)."""
    if not text:
        return False
    text = str(text).strip()
    return text.startswith("กำไรต่อหุ้น") or "กำไรต่อหุ้นขั้นพื้นฐาน" in text


def _detect_period_months(ws) -> int:
    """Infer the reporting-period length (in months) from the sheet's
    top header band.

    Returns one of ``3`` | ``6`` | ``9`` | ``12``, or ``0`` if none of
    the standard SET period phrases appear. Used to pick the
    "cumulative" PL sheet out of the pair that Q1/H1/9M filings ship
    (3-month standalone + longer-period cumulative).
    """
    # Order: longest phrase first. "เก้า" is checked before "สาม"
    # because some sheets state both phrases (prior-period comparative)
    # and we want the longer (cumulative) period to win.
    # SET filers use "สำหรับงวด..." and "สำหรับรอบระยะเวลา..."
    # interchangeably; both are listed as keys.
    patterns = [
        (12, ("สำหรับปี", "สำหรับรอบปี", "สำหรับงวดสิบสองเดือน",
              "สำหรับรอบระยะเวลาหนึ่งปี", "สิบสองเดือน", "(12M)")),
        (9,  ("เก้าเดือน", "9 เดือน",
              "สำหรับงวดเก้าเดือน", "สำหรับรอบระยะเวลาเก้าเดือน",
              "(9M)")),
        (6,  ("หกเดือน", "6 เดือน",
              "สำหรับงวดหกเดือน", "สำหรับรอบระยะเวลาหกเดือน",
              "(6M)")),
        (3,  ("สามเดือน", "3 เดือน",
              "สำหรับงวดสามเดือน", "สำหรับรอบระยะเวลาสามเดือน",
              "(3M)")),
    ]
    max_row = min(14, ws.max_row)
    for months, keys in patterns:
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    s = str(cell)
                    if any(k in s for k in keys):
                        return months
    return 0


def _find_period_transition(ws) -> Optional[int]:
    """Some filers (KTB, LHFG, BAY-style) pack BOTH the 3-month
    standalone and the N-month cumulative sections into a single PL
    sheet, stacked vertically. Return the row index where the period
    header first changes from a short period to a longer one — that's
    the start of the cumulative section.

    Returns ``None`` when the sheet sticks to a single period throughout.
    """
    transitions: list[tuple[int, int]] = []   # (row, months)
    patterns = [
        (12, ("สำหรับปี", "สำหรับรอบปี",
              "สำหรับรอบระยะเวลาหนึ่งปี", "สำหรับงวดสิบสองเดือน",
              "สิบสองเดือน")),
        (9,  ("เก้าเดือน", "9 เดือน",
              "สำหรับงวดเก้าเดือน", "สำหรับรอบระยะเวลาเก้าเดือน")),
        (6,  ("หกเดือน", "6 เดือน",
              "สำหรับงวดหกเดือน", "สำหรับรอบระยะเวลาหกเดือน")),
        (3,  ("สามเดือน", "3 เดือน",
              "สำหรับงวดสามเดือน", "สำหรับรอบระยะเวลาสามเดือน")),
    ]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        for cell in row:
            if cell and isinstance(cell, str):
                s = str(cell)
                for months, keys in patterns:
                    if any(k in s for k in keys):
                        if not transitions or transitions[-1][1] != months:
                            transitions.append((i, months))
                        break
                break
    # First transition row where the period jumps to a longer duration.
    for idx in range(1, len(transitions)):
        prev_months = transitions[idx - 1][1]
        cur_months = transitions[idx][1]
        if cur_months > prev_months:
            return transitions[idx][0]
    return None


def _extract_shareholder_from_rows(ws, start_row: int, end_row: int,
                                    unit_divisor: float) -> tuple[Optional[float], Optional[float]]:
    """Walk rows ``[start_row, end_row]`` and return the first
    (shareholder_profit, prior) pair found. Falls back to the top-line
    net-profit row if the shareholder split isn't broken out in that
    section. Values come back divided by ``unit_divisor`` (MB)."""
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        if not row:
            continue
        label, label_col = _find_label(row)
        if not label:
            continue
        if _is_shareholder_profit_row(label):
            nums = _extract_numeric(row, is_eps=False, start=label_col + 1)
            if len(nums) >= 2:
                return nums[0] / unit_divisor, nums[1] / unit_divisor
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        if not row:
            continue
        label, label_col = _find_label(row)
        if not label:
            continue
        if _is_netprofit_row(label):
            nums = _extract_numeric(row, is_eps=False, start=label_col + 1)
            if len(nums) >= 2:
                return nums[0] / unit_divisor, nums[1] / unit_divisor
    return None, None


def _find_pl_sheets(workbook) -> list[str]:
    """Return every Profit & Loss sheet in the workbook, ordered from
    shortest reporting period (standalone) to longest (cumulative).

    Q1/H1/9M filings ship two PL sheets — a 3-month standalone plus a
    6/9-month cumulative — and we need both so we can derive Q2 and Q4
    for issuers that skip the H1 filing (most commercial banks).

    Selection is name-based first (``PL``, ``กำไรขาดทุน``, or numeric
    names like ``'8-9'`` / ``'PL3M-6-7'``), with a content-based
    fallback scanning for ``กำไรสุทธิ`` on sheets that don't carry an
    obvious label (e.g. KKP's ``'8-9'``).
    """
    # Filers occasionally leave abandoned/template PL sheets in the workbook
    # and label them "ไม่เอา" / "ไม่ใช้" / "ไม่ใช้แล้ว" / "ห้ามใช้" / "DO NOT USE"
    # — literally telling the reader to skip them. TEAM 2564 FY ships a
    # 'PL-ไม่เอา' sheet with stale 2561 quarterly numbers next to the real
    # 'PL' annual sheet; without this filter the parser picks the wrong
    # sheet as primary and derives a 2561 quarterly from a 2564 FY filing.
    def _is_skip_sheet(name: str) -> bool:
        n = name.replace(" ", "").upper()
        # Thai disposal markers — use the no-space variant to catch both
        # "ไม่เอา" and "ไม่ ใช้" formatting.
        skip_markers_th = ("ไม่เอา", "ไม่ใช้", "ห้ามใช้", "ห้ามอ่าน", "ยกเลิก")
        if any(m in name for m in skip_markers_th):
            return True
        # English equivalents.
        return any(m in n for m in ("DONOTUSE", "DO_NOT_USE", "DEPRECATED"))

    candidates: list[str] = []
    for name in workbook.sheetnames:
        if _is_skip_sheet(name):
            continue
        upper = name.upper()
        if "PL" in upper or "กำไรขาดทุน" in name:
            candidates.append(name)

    # Content fallback — any sheet whose body contains an
    # income-statement marker. Always run alongside the name-match pass
    # (not just when name-match is empty): some filers (FE) ship a 2568
    # FY workbook where the only "PL"-named sheets are stale leftovers
    # from 2555-2559 and the current PL is buried inside a sheet named
    # 'งบการเงิน'. Without merging the content-detected sheet, the
    # parser locks onto the stale data.
    markers = (
        "งบกำไรขาดทุน",          # PL header — most reliable single marker
        "กำไรสุทธิ",
        "กำไรสำหรับงวด",
        "กำไรสำหรับปี",
        "กำไรสำหรับไตรมาส",
        "กำไรสำหรับรอบระยะเวลา",
        "ส่วนที่เป็นของผู้ถือหุ้น",
        "ส่วนที่เป็นของบริษัท",
        "ส่วนที่เป็นของธนาคาร",
    )
    for name in workbook.sheetnames:
        if _is_skip_sheet(name) or name in candidates:
            continue
        ws = workbook[name]
        max_row = min(160, ws.max_row)
        found = False
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    s = str(cell)
                    if any(m in s for m in markers):
                        found = True
                        break
            if found:
                break
        if found:
            candidates.append(name)

    # Drop empty templates: filers occasionally ship a placeholder sheet
    # ("Sheet1") with the PL skeleton but no numbers — CHOTI 2567 FY is
    # one example. A real PL sheet has at least one numeric value paired
    # with a net-profit-style label. Walking in iter_rows is cheap.
    def _has_pl_data(name: str) -> bool:
        """Return True iff the sheet has at least one populated
        net-profit / shareholder-profit row. Strict matcher only —
        comprehensive-income-only sheets are intentionally rejected
        because picking them as primary would shadow sibling sheets
        that carry the proper ``กำไรสุทธิสำหรับปี`` line (THREL ships
        both layouts and the strict sheet is the right one)."""
        ws = workbook[name]
        max_row = min(ws.max_row, 200)
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
            if not row:
                continue
            label, label_col = _find_label(row)
            if not label:
                continue
            if not (_is_netprofit_row(label) or _is_shareholder_profit_row(label)):
                continue
            end = min(len(row), 30)
            for cell in row[label_col + 1:end]:
                if isinstance(cell, (int, float)) and cell not in (0, None):
                    return True
        return False

    # Stale-sheet filter: some filers (TEAM 2564 FY) ship workbooks with
    # leftover PL sheets from earlier filings — e.g. a 2561 quarterly
    # sheet stuck in a 2564 FY zip. Without filtering, the period-length
    # sort below picks the stale 3-month sheet as primary because it's
    # "shorter" than the real 12-month FY sheet. Drop any candidate
    # whose latest year (parsed from its top header band) is older than
    # the latest year present anywhere in the candidate set.
    #
    # Run BEFORE the data-tier filter — otherwise BEAUTY 2566 FY (whose
    # only "real net-profit" candidate is the stale 'PL  (งวดสามเดือน)'
    # sheet from 2562) would lock onto the wrong sheet, with the actual
    # current-year 'PL' sheet dropped at the tier step before year
    # comparisons run.
    def _latest_year_in_sheet(name: str) -> int:
        ws = workbook[name]
        latest = 0
        for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row),
                                values_only=True):
            # Year header rows contain only year-shaped integers (or
            # strings); data rows contain money figures alongside their
            # Thai labels. Skip rows that look like data — otherwise a
            # money figure whose value happens to fall in [2540, 2600]
            # (e.g. SCB 2567 Q1 cash-flow row has 2,589 K baht in the
            # depreciation column) gets mistaken for a Buddhist year and
            # poisons the workbook_latest, which then filters out the
            # real PL sheet as "stale".
            has_money = any(
                isinstance(c, (int, float)) and abs(c) > 9999
                for c in row
            )
            if has_money:
                continue
            for cell in row:
                if cell is None:
                    continue
                # Year as integer cell (column header row).
                if isinstance(cell, (int, float)):
                    iv = int(cell)
                    if 2540 <= iv <= 2600 and iv > latest:
                        latest = iv
                # Year embedded in date string (e.g. "31 ธันวาคม 2564").
                elif isinstance(cell, str):
                    for m in re.findall(r"25\d{2}", cell):
                        iv = int(m)
                        if 2540 <= iv <= 2600 and iv > latest:
                            latest = iv
        return latest

    sheet_years = {n: _latest_year_in_sheet(n) for n in candidates}
    workbook_latest = max(sheet_years.values(), default=0)
    if workbook_latest:
        candidates = [
            n for n in candidates
            if sheet_years[n] == 0 or sheet_years[n] >= workbook_latest
        ]

    candidates = [c for c in candidates if _has_pl_data(c)]

    # Order by period length so [0] is the 3-month standalone sheet and
    # [-1] is the longest cumulative. Sheets with an unknown period sit
    # at the end — they're rare and won't affect 3m/cum pairing.
    candidates.sort(key=lambda n: _detect_period_months(workbook[n]) or 99)
    return candidates


def _find_pl_sheet(workbook) -> Optional[str]:
    """Back-compat wrapper — returns the primary (shortest-period) PL
    sheet, matching the original single-return contract."""
    sheets = _find_pl_sheets(workbook)
    return sheets[0] if sheets else None


def _detect_unit_divisor(rows_top: list) -> float:
    """Infer how to convert XLSX cell values to millions of baht.

    SET financial XLSX files include a unit marker near the top of every
    sheet (e.g. ``(บาท)`` for annual reports, ``(พันบาท)`` for quarterly,
    occasionally ``(ล้านบาท)`` for summarised sheets).

    Returns the divisor to apply to raw numeric cells so the result is in
    millions of baht. Defaults to 1,000,000 (treat as baht) if no marker
    is found — the original behaviour for XLSX where the unit was implicit.
    """
    for row in rows_top:
        for cell in row:
            if not cell or not isinstance(cell, str):
                continue
            s = cell.strip()
            # Order matters: check more-specific markers first.
            if "ล้านบาท" in s:
                return 1.0
            if "พันบาท" in s:
                return 1_000.0
            if "บาท" in s:
                return 1_000_000.0
    return 1_000_000.0


def _build_unit_divisor_map(ws, max_rows: int = 200) -> dict[int, float]:
    """Map row index → divisor for sheets that stack multiple period
    blocks with DIFFERENT units.

    BTW / QDC / PLANET / TNITY ship a single PL sheet that stacks the
    current-quarter block (unit ``พันบาท``) on top of the annual block
    (unit ``บาท``). A single top-of-sheet detection picks ``พันบาท`` and
    applies it to the FY rows below — yielding values 1,000× too large
    for the annual figure SET publishes.

    The map records the divisor implied by every unit marker seen in
    the sheet, keyed by the row where the marker appeared. Lookup at
    extraction time picks the most recent marker AT-OR-ABOVE the
    target row, so each block gets its own correct divisor.
    """
    out: dict[int, float] = {}
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                     values_only=True)
    ):
        for cell in row:
            if not cell or not isinstance(cell, str):
                continue
            s = cell.strip()
            if "ล้านบาท" in s:
                out[i] = 1.0
                break
            if "พันบาท" in s:
                out[i] = 1_000.0
                break
            if "บาท" in s:
                out[i] = 1_000_000.0
                break
    return out


def _divisor_for_row(unit_map: dict[int, float], row_idx: int,
                     default: float) -> float:
    """Return the divisor that applies to ``row_idx`` — the most recent
    marker at or above the row. Falls back to ``default`` (the
    sheet-wide value detected from the top band) when nothing earlier
    matches."""
    candidates = [r for r in unit_map.keys() if r <= row_idx]
    if not candidates:
        return default
    return unit_map[max(candidates)]


def _detect_year_columns(rows_top: list) -> set[int]:
    """Return the set of column indices that the sheet uses as year
    headers in its top band.

    Scans the first ~15 rows for cells whose value parses as a
    Buddhist Era year integer (2540-2600) — either as a numeric cell
    or as a string like ``'2566'`` / ``'2566 (ปรับปรุงใหม่)'``. Skips
    rows that also contain a money-shaped number (``abs > 9999``) so
    a stray depreciation figure doesn't poison the set (same guard
    used by ``_latest_year_in_sheet``).

    The result is the canonical "data columns" for `_extract_numeric` —
    everything outside is either a label column, a note reference, or
    an orphan column like BTC's share-count column at col 4.
    """
    cols: set[int] = set()
    for row in rows_top:
        has_money = any(
            isinstance(c, (int, float)) and abs(c) > 9999
            for c in row
        )
        if has_money:
            continue
        for j, cell in enumerate(row):
            if cell is None:
                continue
            if isinstance(cell, (int, float)):
                iv = int(cell)
                if 2540 <= iv <= 2600:
                    cols.add(j)
            elif isinstance(cell, str):
                m = re.search(r"25\d{2}", cell)
                if m:
                    iv = int(m.group(0))
                    if 2540 <= iv <= 2600:
                        cols.add(j)
    return cols


def _detect_thb_column_offset(rows_top: list) -> int:
    """Return the column index where THB data starts in a dual-currency
    layout, or 0 if the sheet is single-currency (so the default
    ``label_col + 1`` start applies).

    PTTEP / BANPU / SPRC ship workbooks with USD and THB side-by-side:
    the unit row reads ``หน่วย: พันเหรียญสหรัฐ`` over the first numeric
    block and ``หน่วย: พันบาท`` over a second block further right. SET's
    company-highlight figures are THB, so we have to skip the USD
    columns. We detect dual-currency by seeing both markers in the same
    top band, and return the column of the THB marker.
    """
    has_usd = False
    has_thb_marker = False
    usd_max_col = -1
    for row in rows_top:
        for col, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            if "ดอลลาร์" in cell or "เหรียญสหรัฐ" in cell:
                has_usd = True
                if col > usd_max_col:
                    usd_max_col = col
            # SPRC 2566+ uses bare ``บาท`` cells in the unit row (one
            # per data column) instead of a single merged ``หน่วย: บาท``.
            # Recognise the bare form too so we don't miss the THB
            # block — but only when the cell IS the word "บาท" (stripped),
            # otherwise we'd pick up running text that happens to mention
            # the currency (e.g. ``กำไรต่อหุ้น (บาท)``).
            stripped = cell.strip()
            if (
                "พันบาท" in cell
                or "ล้านบาท" in cell
                or ("หน่วย" in cell and "บาท" in cell)
                or stripped == "บาท"
            ):
                has_thb_marker = True

    if not (has_usd and has_thb_marker):
        return 0

    # The year-header row carries one ``25xx`` cell per data column.
    # Dual-currency layouts come in two shapes:
    #   • Simple 4-col (BANPU/PTTEP/CCET): USD CONSO + USD prior +
    #     THB CONSO + THB prior. Latest year appears twice — take the
    #     SECOND occurrence as the THB column.
    #   • Complex 6-8 col (SPRC 2566+): USD CONSO + USD SEP + USD prior
    #     + THB CONSO + THB SEP + THB prior. Latest year appears 4
    #     times — the second occurrence is USD SEP (still wrong); we
    #     need the FIRST occurrence whose column is past the USD block.
    # Use ``usd_max_col`` (max position of any USD marker) as the
    # boundary: the THB current-period column is the first latest-year
    # column to its right.
    latest_year = 0
    for row in rows_top:
        for cell in row:
            if isinstance(cell, (int, float)):
                iv = int(cell)
                if 2540 <= iv <= 2600 and iv > latest_year:
                    latest_year = iv
            elif isinstance(cell, str):
                for m in re.findall(r"25\d{2}", cell):
                    iv = int(m)
                    if 2540 <= iv <= 2600 and iv > latest_year:
                        latest_year = iv
    if latest_year == 0:
        return 0

    target_full = str(latest_year)
    occurrences: list[int] = []
    for row in rows_top:
        per_row: list[int] = []
        for col, cell in enumerate(row):
            matched = False
            if isinstance(cell, (int, float)) and int(cell) == latest_year:
                matched = True
            elif isinstance(cell, str):
                stripped = cell.strip()
                # Year-HEADER cells are short ("2567" or "พ.ศ. 2567",
                # ≤ ~20 chars). Title or date strings ("สำหรับปีสิ้นสุด
                # วันที่ 31 ธันวาคม พ.ศ. 2567") also mention the year
                # but aren't column headers; filtering by length keeps
                # them out of the occurrence set.
                if len(stripped) <= 20 and target_full in stripped:
                    matched = True
            if matched:
                per_row.append(col)
        # Prefer the row that actually carries multiple year columns
        # (the header row), not isolated mentions in title strings.
        if len(per_row) >= 2:
            occurrences = per_row
            break

    if not occurrences:
        return 0

    # First occurrence past the USD block — handles SPRC's 4-section
    # layout cleanly.
    if usd_max_col >= 0:
        past_usd = [c for c in occurrences if c > usd_max_col]
        if past_usd:
            return past_usd[0]

    # Fallback: take the second occurrence (BANPU/CCET/PTTEP shape).
    if len(occurrences) >= 2:
        return occurrences[1]
    return 0


class _XlsAdapter:
    """Minimal subset of openpyxl's API backed by xlrd, so the rest of the
    parser can treat .xls and .xlsx identically."""

    class Sheet:
        def __init__(self, sheet):
            self._s = sheet
            self.max_row = sheet.nrows
            self.max_column = sheet.ncols

        def iter_rows(self, min_row=1, max_row=None, values_only=True):
            end = self._s.nrows if max_row is None else min(max_row, self._s.nrows)
            for r in range(min_row - 1, end):
                yield tuple(self._s.row_values(r))

    def __init__(self, xls_path: str):
        self._book = xlrd.open_workbook(xls_path)
        self.sheetnames = self._book.sheet_names()

    def __getitem__(self, name: str) -> "Sheet":
        return self.Sheet(self._book.sheet_by_name(name))


def _open_workbook(path: str):
    """Open .xls or .xlsx and return an object with sheetnames + __getitem__.

    Detect the actual format by magic bytes rather than extension —
    some filers (DELTA, notably) ship xlsx files renamed to ``.xls``.
    We pass the bytes through ``BytesIO`` so openpyxl never sees the
    misleading filename and falls back to its extension check."""
    from io import BytesIO

    with open(path, "rb") as f:
        data = f.read()
    head = data[:8]

    # xlsx / xlsm / etc. — Office Open XML is a ZIP archive.
    if head[:4] == b"PK\x03\x04":
        return openpyxl.load_workbook(BytesIO(data), data_only=True)

    # Legacy BIFF binary .xls
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return _XlsAdapter(path)

    # Last-resort: trust the extension when magic bytes don't match
    # (rarely-used XML SpreadsheetML 2003, etc.).
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return openpyxl.load_workbook(BytesIO(data), data_only=True)
    if lower.endswith(".xls"):
        return _XlsAdapter(path)
    raise ValueError(f"Unsupported workbook format: {path}")


_TAS_REF_PATTERN = re.compile(
    r"^(TAS|TFRS|IAS|IFRS)\b[\s.\d,ก-๙]*$",
    re.IGNORECASE,
)


def _normalize_thai_sara_am(s: str) -> str:
    """Some filers (e.g. MOSHI) write Thai SARA AM as the decomposed
    sequence NIKHAHIT (U+0E4D) + SARA AA (U+0E32) instead of the
    composed SARA AM (U+0E33). They render identically but compare
    unequal as bytes, so ``กำไรสุทธิสำหรับปี`` (real net-profit row)
    fails our regex matchers. Re-compose the pair so all downstream
    label regexes match either form."""
    if "ํา" in s:
        return s.replace("ํา", "ำ")
    return s


def _find_label(row: tuple) -> tuple[str, int]:
    """Find the first meaningful non-empty string cell in ``row`` and
    return ``(label, column_index)``.

    Bank XLSX filings indent the parent-vs-minority split into column B
    or C (e.g. KBANK uses col 1, BBL uses col 2), leaving col A empty.
    We therefore can't hard-code ``row[0]`` as the label.

    MTC and similar filers use column A as a Thai/IFRS reference code
    column (``TAS 1.81ก.1``, ``TFRS 7.23.3``) and put the actual line
    item label in column B. Skipping the reference code lets the
    label-match regexes see ``กำไรสำหรับปี`` instead of the IFRS code.
    """
    for i, cell in enumerate(row):
        if cell is None:
            continue
        if not isinstance(cell, str):
            continue
        s = _normalize_thai_sara_am(str(cell).strip())
        if not s:
            continue
        if _TAS_REF_PATTERN.match(s):
            continue
        return s, i
    return "", -1


def _extract_numeric(row: tuple, is_eps: bool = False, start: int = 1,
                     year_cols: Optional[set[int]] = None) -> list:
    """Extract first 2 non-zero numeric values from row (current, prior).
    Skips small integers (1-99) which are footnote references in SET XBRL.

    For non-EPS: skips whole-number values < 100 (likely notes)
    For EPS: skips whole integers (which are notes), accepts decimals

    ``start`` is the first column index to scan — callers set this to one
    past the label column so the label itself is never misread as a value.

    ``year_cols`` (optional) restricts extraction to a known set of
    year-header column indices detected from the top band. BTC's PL
    sheet, for instance, has an orphan share-count column (col 4 =
    852,812,933 = shares outstanding) immediately before the real
    year-data columns (5/7/9/11). Without the whitelist the parser
    picks the share count as "current year" and reports the same wrong
    number on every row of every BTC year.

    Hard cap the scan at column 20 (relative to row start). Real SET PL
    sheets keep their current/prior data within the first ~15 columns;
    anything beyond that is helper buckets, internal cross-checks, or
    stale cached values from broken formulas. DITTO 2566/Q1, for
    example, has #REF! in cols 7–13 (the real data columns) and stale
    FY 2565 totals at col 31 — without the cap we'd silently pick up
    the stale FY value as Q1 net profit, blowing up Q-sums by 1000×.
    """
    end = min(len(row), 30)
    # Pre-scan all numerics in range so we can apply lookahead heuristics
    # (footnote references that aren't < 100 — e.g. ERW PL-9 has
    # 'ส่วนที่เป็นของบริษัทใหญ่' rows with note ref ``3935`` followed by
    # the actual 838M data column).
    raw_cells: list[tuple[int, float, bool]] = []  # (index, val, is_int_type)
    for i, cell in enumerate(row[start:end], start):
        if not isinstance(cell, (int, float)) or cell == 0:
            continue
        # Year-column whitelist: when we know which columns the top
        # header marked as year columns, drop cells in any other
        # column. Falls through when year_cols is None or empty so
        # legacy callers and weird sheets without year integers in
        # the header still get the original behaviour.
        if year_cols and i not in year_cols:
            continue
        raw_cells.append((i, float(cell), isinstance(cell, int)))

    nums = []
    for k, (_idx, val, is_int_type) in enumerate(raw_cells):
        if is_eps:
            # EPS is always a decimal. Notes are whole integers. Skip any
            # integer value or whole-number float.
            if is_int_type:
                continue
            if val == int(val):
                continue
        else:
            # Standard footnote skip: small whole integers / short decimals.
            if abs(val) < 100:
                continue
            # Magnitude lookahead: a real data value sits next to OTHER
            # data values of similar magnitude. If the very next non-zero
            # numeric in this row is dramatically larger AND the current
            # value is small enough that it could plausibly be a note ref
            # (< 10,000 — typical footnote refs are 1-9999), treat the
            # current cell as a footnote and skip. Two guards:
            #   1. ratio > 1000 — real data within a single row rarely
            #      varies by more than ~100x even for volatile sectors.
            #      ERW had 3,935 followed by 838M = ratio 213,000x.
            #   2. abs(val) < 10,000 — keeps SCC 2566 Q1's
            #      ``กำไรสำหรับงวด`` 55,307 thousand baht (real value),
            #      where the prior column 8,756,158 was ~158x bigger
            #      (legitimate Q-over-Q dividend swing, not a footnote).
            if (
                k + 1 < len(raw_cells)
                and abs(val) < 10_000
            ):
                next_val = raw_cells[k + 1][1]
                if abs(next_val) > abs(val) * 1000:
                    continue
        nums.append(val)
        if len(nums) >= 2:
            break
    return nums


def _parse_filename(filename: str) -> Dict[str, Any]:
    """Parse SET zip filename.
    Format: {code}FIN{DD}{MM}{YYYY}{HHMMSS}{seq}T.zip
    Example: 0737FIN250220261406350902T.zip
    """
    m = re.match(r"(\d{4})(FIN|ANN|NWS|F56)(\d{2})(\d{2})(\d{4})(\d{6})(\d{4})T\.zip",
                 filename, re.IGNORECASE)
    if m:
        code, rtype, dd, mm, yyyy, _, _ = m.groups()
        return {
            "company_code": code,
            "report_type": rtype,
            "report_date": f"{yyyy}-{mm}-{dd}",
        }
    return {}


def parse_zip(zip_path: str, symbol: str = "UNKNOWN") -> Optional[FinancialData]:
    """Parse a SET news zip file and extract key financial data.

    Args:
        zip_path: Path to the zip file
        symbol: Stock symbol (e.g. 'CPALL')

    Returns:
        FinancialData object, or None if parsing failed.
    """
    if not os.path.exists(zip_path):
        raise FileNotFoundError(f"Zip not found: {zip_path}")

    filename = os.path.basename(zip_path)
    meta = _parse_filename(filename)

    with tempfile.TemporaryDirectory() as tmp:
        # Extract zip
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(tmp)

        # Find a financial-statement workbook (.xlsx preferred, fall back to .xls)
        xlsx_path = xls_path = None
        for root, _, files in os.walk(tmp):
            for f in files:
                upper = f.upper()
                if upper.endswith(".XLSX") and not xlsx_path:
                    xlsx_path = os.path.join(root, f)
                elif upper.endswith(".XLS") and not xls_path:
                    xls_path = os.path.join(root, f)

        book_path = xlsx_path or xls_path
        if not book_path:
            print(f"[parse_zip] No XLS/XLSX found in {filename}")
            return None

        # Open workbook (unified adapter handles both formats)
        try:
            wb = _open_workbook(book_path)
        except Exception as e:
            print(f"[parse_zip] Failed to open workbook: {e}")
            return None

        pl_sheets = _find_pl_sheets(wb)
        if not pl_sheets:
            print(f"[parse_zip] No PL sheet in {filename}")
            return None

        pl_sheet = pl_sheets[0]
        ws = wb[pl_sheet]

        # Detect the unit divisor from the top of the sheet
        top_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        unit_divisor = _detect_unit_divisor(top_rows)
        # Stacked-block sheets (BTW/QDC/PLANET/TNITY) carry one PL block
        # per period with DIFFERENT units (พันบาท over the quarterly
        # block, บาท over the annual block). Build a row→divisor map so
        # the FY block below uses its own unit, not the top one.
        unit_map = _build_unit_divisor_map(ws)
        # Year-column whitelist: only extract numerics from columns the
        # top-band header marked as year columns. Without this, BTC's
        # orphan share-count column (col 4 = 852,812,933) silently
        # masquerades as a current-year profit.
        year_cols = _detect_year_columns(top_rows)
        # Dual-currency offset: PTTEP/BANPU lay USD columns first, THB
        # columns after — start the per-row scan past the USD columns so
        # extraction picks up the THB values SET reports.
        thb_offset = _detect_thb_column_offset(top_rows)

        # Detect period type from first few rows
        period_label = ""
        period_type = "unknown"
        year = 0
        quarter = None

        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text = str(cell).strip()
                    # Period type detection
                    if ("สำหรับปี" in text or "สำหรับรอบปี" in text
                            or "สำหรับรอบระยะเวลาหนึ่งปี" in text):
                        period_type = "annual"
                    elif ("สำหรับงวดหกเดือน" in text or "6 เดือน" in text
                          or "หกเดือน" in text
                          or "สำหรับรอบระยะเวลาหกเดือน" in text):
                        period_type = "half"
                    elif ("สำหรับไตรมาส" in text or "3 เดือน" in text
                          or "สำหรับงวดสามเดือน" in text
                          or "สำหรับรอบระยะเวลาสามเดือน" in text):
                        period_type = "quarterly"

                    # Year from Thai date text — keep the LATEST so we
                    # don't get tricked by the prior-period column header
                    # (KBANK 2568 FY has '2567 (ปรับปรุงใหม่)' that beat
                    # the integer cell '2568' to the punch under the
                    # original first-match logic).
                    for m in re.findall(r"25\d{2}", text):
                        cand = int(m)
                        if 2540 <= cand <= 2600 and cand > year:
                            year = cand

        # Also sweep header row cells for integer year values; pick the
        # MAX across the whole header band, not the first one seen.
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, (int, float)):
                    val = int(cell)
                    if 2540 <= val <= 2600 and val > year:
                        year = val

        # Fallback to filename date (Gregorian year - 543 = Thai year)
        if year == 0 and "report_date" in meta:
            try:
                gregorian = int(meta["report_date"][:4])
                year = gregorian - 543 - 1  # Report filed in following year
            except ValueError:
                pass

        # Extract financial data
        result = FinancialData(
            symbol=symbol,
            filename=filename,
            period_label=period_label or f"FY{year}" if period_type == "annual" else "",
            period_type=period_type,
            year=year,
            quarter=quarter,
        )

        # ITC 2564 / ANAN 2564 split the parent-share line by "continuing"
        # vs "discontinued" operations and emit a section-total row at
        # the end. Two layouts seen:
        #   ITC: ``รวม`` row label + values
        #   ANAN: blank label, just values (label visually merged in
        #         Excel but openpyxl reads no string)
        # Track when we last saw a parent-share section header without
        # values; the next row carrying values inside that section is
        # the parent total. Stop tracking on NCI / minority headers so
        # we don't grab the NCI total instead.
        # Stacked-block sheets (BTW, QDC, PLANET) put a quarterly block
        # on top (พันบาท, divisor 1,000) and the annual block below
        # (บาท, divisor 1,000,000). Without this guard the parser locks
        # onto the first shareholder row in the quarterly block — wrong
        # period AND wrong unit. We let a later, *higher-precision* block
        # (larger divisor) override; same-or-lower divisor keeps the
        # original "first wins" semantics so simple single-block sheets
        # are unaffected.
        last_sh_divisor = 0.0    # divisor used when shareholder was set
        last_np_divisor = 0.0    # divisor used when net_profit was set
        in_parent_section = False
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
        ):
            if not row:
                continue
            # Recompute the unit divisor from the most recent marker
            # at-or-above this row. Static `unit_divisor` is the
            # fall-through default for sheets without per-block markers.
            row_divisor = _divisor_for_row(unit_map, row_idx, unit_divisor)
            label, label_col = _find_label(row)

            # NCI section closes any open parent section. Even on an
            # unlabeled row we'd still want to know we exited.
            if label and ("ส่วนได้เสีย" in label or "ไม่มีอำนาจ" in label):
                in_parent_section = False
                continue

            if not label:
                # Unlabeled numeric row — only meaningful as the parent
                # total when we're already inside a parent section.
                if in_parent_section and result.shareholder_profit is None:
                    scan_start = thb_offset if thb_offset else 0
                    nums = _extract_numeric(row, is_eps=False, start=scan_start,
                                            year_cols=year_cols)
                    if len(nums) >= 2:
                        result.shareholder_profit = nums[0] / row_divisor
                        result.shareholder_profit_prior = nums[1] / row_divisor
                        in_parent_section = False
                continue

            is_eps = _is_eps_row(label)
            # Dual-currency layouts skip past the USD block; otherwise
            # start scanning at the column right after the label.
            scan_start = max(label_col + 1, thb_offset) if thb_offset else label_col + 1
            nums = _extract_numeric(row, is_eps=is_eps, start=scan_start,
                                    year_cols=year_cols)

            stripped = label.strip().rstrip(":")

            if len(nums) >= 2:
                if _is_revenue_row(label) and result.revenue is None:
                    result.revenue = nums[0] / row_divisor
                    result.revenue_prior = nums[1] / row_divisor
                elif _is_netprofit_row(label) and (
                    result.net_profit is None or row_divisor > last_np_divisor
                ):
                    result.net_profit = nums[0] / row_divisor
                    result.net_profit_prior = nums[1] / row_divisor
                    last_np_divisor = row_divisor
                elif _is_shareholder_profit_row(label) and (
                    result.shareholder_profit is None or row_divisor > last_sh_divisor
                ):
                    result.shareholder_profit = nums[0] / row_divisor
                    result.shareholder_profit_prior = nums[1] / row_divisor
                    last_sh_divisor = row_divisor
                    in_parent_section = False
                elif is_eps and result.eps is None:
                    # EPS is always in baht per share regardless of sheet unit
                    result.eps = nums[0]
                    result.eps_prior = nums[1]
                elif (
                    in_parent_section
                    and result.shareholder_profit is None
                    and stripped == "รวม"
                ):
                    # Total of the parent-share split (ITC 2564).
                    result.shareholder_profit = nums[0] / row_divisor
                    result.shareholder_profit_prior = nums[1] / row_divisor
                    in_parent_section = False
            else:
                # Header-only row (no values). Mark when this is a
                # parent-share section header — the next row with
                # values is the parent total.
                if (
                    _is_shareholder_profit_row(label)
                    and result.shareholder_profit is None
                ):
                    in_parent_section = True

        # Cross-sheet recovery: some filers (2S FY2568, others that split
        # the income statement across two pages) put only the consolidated
        # total on the primary PL sheet and stash the parent-vs-NCI
        # breakdown on a secondary sheet. Sweep the remaining PL sheets
        # for a parent-share row before falling back to net_profit.
        # SET's company-highlight API reports the parent-share figure, so
        # using the consolidated total here would silently mismatch by
        # the NCI amount (e.g. 2S FY2568: 144.35 total vs 144.53 parent).
        #
        # The continuation "(ต่อ)" sheets typically omit the หมายเหตุ
        # footnote column — the first cell after the label is the current
        # period's value. _extract_numeric's abs<100 footnote-skip
        # heuristic would silently drop small parent-share values
        # (2S Q3/2568 parent = -30 พันบาท), so we use a direct
        # first-two-nonzero scan here instead.
        if result.shareholder_profit is None and len(pl_sheets) > 1:
            for cand in pl_sheets[1:]:
                ws_c = wb[cand]
                cu = _detect_unit_divisor(
                    list(ws_c.iter_rows(min_row=1, max_row=15, values_only=True))
                )
                found = False
                for row in ws_c.iter_rows(min_row=1, max_row=ws_c.max_row, values_only=True):
                    if not row:
                        continue
                    label, lcol = _find_label(row)
                    if not label or not _is_shareholder_profit_row(label):
                        continue
                    nums: list[float] = []
                    for cell in row[lcol + 1: min(len(row), 20)]:
                        if isinstance(cell, (int, float)) and cell != 0:
                            nums.append(float(cell))
                            if len(nums) >= 2:
                                break
                    if len(nums) >= 2:
                        result.shareholder_profit = nums[0] / cu
                        result.shareholder_profit_prior = nums[1] / cu
                        found = True
                        break
                if found:
                    break

        # Bank filings report a single consolidated "กำไรสุทธิ" but never
        # break it into shareholder/minority on some pages — fall back to
        # the top-line net profit so the symbol isn't silently dropped.
        if result.shareholder_profit is None and result.net_profit is not None:
            result.shareholder_profit = result.net_profit
            result.shareholder_profit_prior = result.net_profit_prior

        # Record the months covered by the primary sheet so single-sheet
        # Q1 / FY filings still report a sensible "cumulative" (same as
        # the standalone — cumulative == current period for those).
        primary_months = _detect_period_months(ws)
        result.primary_months = primary_months

        # Extract the cumulative-period shareholder profit. Three
        # possible layouts:
        #   1. Separate second PL sheet (CPALL, KBANK, SCB, BBL, …) —
        #      take shareholder profit from that sheet's first match.
        #   2. Single sheet that stacks 3-month and cumulative sections
        #      vertically (KTB, LHFG) — detect the period-header
        #      transition row and pull shareholder profit from the rows
        #      beyond it.
        #   3. Single sheet covering one period only (Q1, FY) — the
        #      cumulative equals the standalone figure we just parsed.
        # Walk the candidate cumulative sheets in order. We want the
        # FIRST one that's longer than the primary AND actually yields
        # a shareholder/net-profit number — non-PL "อื่น ๆ" (changes
        # in equity, etc.) sometimes mention "กำไรสุทธิ" inline and
        # get picked up by the content-based fallback in find_pl_sheets,
        # so just taking the last sheet is unsafe.
        cum_sheet = None
        cum_sp = None
        cum_sp_prior = None
        cum_unit_used = None
        for cand in pl_sheets:
            if cand == pl_sheet:
                continue
            ws_c = wb[cand]
            months_c = _detect_period_months(ws_c)
            if months_c <= primary_months:
                continue
            cu = _detect_unit_divisor(
                list(ws_c.iter_rows(min_row=1, max_row=15, values_only=True))
            )
            sp_try, sp_prior_try = _extract_shareholder_from_rows(
                ws_c, 1, ws_c.max_row, cu
            )
            if sp_try is not None:
                cum_sheet = cand
                cum_sp = sp_try
                cum_sp_prior = sp_prior_try
                cum_unit_used = cu
                break

        if cum_sheet is not None:
            # Layout 1 — dedicated cumulative sheet.
            result.shareholder_profit_cum = cum_sp
            result.shareholder_profit_cum_prior = cum_sp_prior
            result.cum_months = _detect_period_months(wb[cum_sheet])
        else:
            transition = _find_period_transition(ws)
            if transition is not None:
                # Layout 2 — same sheet, section break at ``transition``.
                # The primary shareholder_profit we already parsed is
                # the earliest match, which is before the transition —
                # i.e. the 3-month standalone. We just need the
                # cumulative from rows after the break.
                #
                # Stacked-block sheets (BTW / QDC / PLANET) put a
                # different unit marker (บาท) at the top of the second
                # block. Detect it from rows AROUND the transition so
                # the cumulative value isn't divided by the first
                # block's marker (พันบาท).
                cum_divisor = _detect_unit_divisor(
                    list(ws.iter_rows(
                        min_row=max(1, transition - 2),
                        max_row=min(ws.max_row, transition + 12),
                        values_only=True,
                    ))
                )
                sp, sp_prior = _extract_shareholder_from_rows(
                    ws, transition, ws.max_row, cum_divisor
                )
                result.shareholder_profit_cum = sp
                result.shareholder_profit_cum_prior = sp_prior
                # Detect cum period by scanning rows past the transition.
                cum_months = 0
                for row in ws.iter_rows(min_row=transition, max_row=min(transition + 10, ws.max_row), values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            s = str(cell)
                            if "เก้าเดือน" in s or "9 เดือน" in s:
                                cum_months = 9; break
                            if "หกเดือน" in s or "6 เดือน" in s:
                                cum_months = 6; break
                            if "สำหรับปี" in s:
                                cum_months = 12; break
                    if cum_months:
                        break
                result.cum_months = cum_months
            else:
                # Layout 3 — truly single-period filing (Q1, FY).
                result.shareholder_profit_cum = result.shareholder_profit
                result.shareholder_profit_cum_prior = result.shareholder_profit_prior
                result.cum_months = primary_months

        # Build period label
        if not result.period_label:
            if period_type == "annual":
                result.period_label = f"งบประจำปี {year}"
            elif period_type == "half":
                result.period_label = f"งบครึ่งปี {year}"
            elif period_type == "quarterly":
                result.period_label = f"งบไตรมาส {year}"
            else:
                result.period_label = f"งบการเงิน {year}"

        return result


def to_dict(data: FinancialData) -> dict:
    """Convert FinancialData to dict for JSON storage."""
    return asdict(data)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python parse_set_zip.py <zip_path> [symbol]")
        sys.exit(1)

    zip_path = sys.argv[1]
    symbol = sys.argv[2] if len(sys.argv) > 2 else "UNKNOWN"

    data = parse_zip(zip_path, symbol)
    if data:
        print("=" * 60)
        print(f"EXTRACTED: {data.symbol}  ·  {data.period_label}")
        print("=" * 60)
        for k, v in asdict(data).items():
            if v is not None:
                if isinstance(v, float):
                    print(f"  {k:30} : {v:>12,.2f}")
                else:
                    print(f"  {k:30} : {v}")
    else:
        print("Failed to parse")
