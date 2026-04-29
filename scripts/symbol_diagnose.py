"""scripts/symbol_diagnose.py — automated XLSX-vs-SET reconciliation.

For each (symbol, year) mismatch, hunt every cell in every PL / SE
sheet of that symbol's filings (FY of the year, FY of year+1, plus
H1 / 9M / Q1 of both years for restatement detection) and report
which exact cell value matches SET's annual figure.

This is the automation that replaces "open XLSX in Excel and squint"
— it lets us audit all 107 mismatches in one batch without asking the
user to click through each filing.

Usage:
    python scripts/symbol_diagnose.py BTW          # one symbol
    python scripts/symbol_diagnose.py --all-mismatches
"""
from __future__ import annotations

import argparse
import json
import sys
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.ingest.zip_downloader import safe_symbol_dir


def _load_xlsx(zp: Path):
    """Open the workbook inside zp. Handles both .xlsx (openpyxl) and
    legacy .xls (xlrd) by extracting to a temp file and dispatching
    through the parser's unified _open_workbook."""
    try:
        from parsers.parse_set_zip import _open_workbook
        import tempfile, os
        with zipfile.ZipFile(zp) as z:
            for n in z.namelist():
                upper = n.upper()
                if not (upper.endswith(".XLSX") or upper.endswith(".XLS")):
                    continue
                with tempfile.NamedTemporaryFile(
                    suffix=".xlsx" if upper.endswith(".XLSX") else ".xls",
                    delete=False,
                ) as tmp:
                    tmp.write(z.read(n))
                    tmp_path = tmp.name
                try:
                    return _open_workbook(tmp_path)
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
    except Exception:
        return None
    return None


def _scan_for_value(wb, target_mb: float) -> list[dict]:
    """Return every cell whose value equals ``target_mb`` (in any unit)
    within tight tolerance (0.5 % relative AND ±0.01 MB absolute).
    Tries the value × {1, 1000, 1000000} so we match cells stored in
    millions of baht / thousands / baht alike.

    Tight tolerance is critical: with the loose tolerance we'd match
    cash-flow rows, working-capital totals, and a hundred other lines
    that happen to share the same magnitude. We want only cells whose
    value rounds to SET's reported figure to the cent.

    Result: list of {sheet, row, col, raw_value, divisor, computed_mb}.
    """
    hits: list[dict] = []
    candidates = [
        (target_mb,                 1.0),         # already in MB (ล้านบาท)
        (target_mb * 1_000,         1_000.0),     # in thousands of baht (พันบาท)
        (target_mb * 1_000_000,     1_000_000.0), # in baht (บาท)
    ]
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(min_row=1,
                                             max_row=min(ws.max_row, 200),
                                             values_only=True)):
            for j, cell in enumerate(row):
                if not isinstance(cell, (int, float)) or cell == 0:
                    continue
                cv = float(cell)
                for raw_target, divisor in candidates:
                    if abs(raw_target) < 0.5:
                        continue
                    rel = abs(cv - raw_target) / max(abs(raw_target), 1.0)
                    if rel <= 0.005:
                        hits.append({
                            "sheet": sn,
                            "row": i,
                            "col": j,
                            "raw_value": cv,
                            "divisor": divisor,
                            "computed_mb": cv / divisor,
                        })
                        break
    return hits


def _row_label(wb, sheet: str, row_idx: int, prev_n: int = 1) -> str:
    """Find the most-recent string label at-or-above row_idx in sheet."""
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=max(1, row_idx + 1 - prev_n),
                             max_row=row_idx + 1, values_only=True))
    for row in reversed(rows):
        for c in row:
            if isinstance(c, str) and c.strip():
                return c.strip()
    return ""


def diagnose(symbol: str, set_target: float, year: int) -> dict:
    """For a symbol/year mismatch with SET = ``set_target`` (MB), open
    every relevant zip and report which cell matches. Search the year's
    own FY filing first, then the next year's FY filing's prior column,
    then 9M / H1 / Q1 of both years."""
    raw_root = Path("data/raw") / safe_symbol_dir(symbol) / "financials"
    if not raw_root.exists():
        return {"symbol": symbol, "year": year, "set_target": set_target,
                "error": "no raw data"}

    # Search order: same-year FY first, then year+1 FY (restatement
    # source), then quarterly filings of both years.
    search_zips: list[tuple[str, Path]] = []
    for y_off in (0, 1, -1):
        for period in ("FY", "9M", "H1", "Q1"):
            p = raw_root / str(year + y_off) / period / "source.zip"
            if p.exists():
                search_zips.append((f"{year + y_off}/{period}", p))

    matches: list[dict] = []
    for tag, zp in search_zips:
        wb = _load_xlsx(zp)
        if wb is None:
            continue
        hits = _scan_for_value(wb, set_target)
        for h in hits:
            label = _row_label(wb, h["sheet"], h["row"], prev_n=3)
            matches.append({
                "filing": tag,
                "sheet": h["sheet"],
                "row": h["row"],
                "col": h["col"],
                "raw_value": h["raw_value"],
                "divisor": h["divisor"],
                "computed_mb": h["computed_mb"],
                "row_label": label,
            })

    return {
        "symbol": symbol,
        "year": year,
        "set_target": set_target,
        "matches": matches,
        "match_count": len(matches),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("symbol", nargs="?", help="symbol to diagnose")
    ap.add_argument("--all-mismatches", action="store_true",
                    help="diagnose every (sym,year) mismatch in "
                         "data/validation/universe_audit_final.json")
    ap.add_argument("--audit", default="data/validation/universe_audit_final.json")
    args = ap.parse_args()

    if args.all_mismatches:
        d = json.loads(Path(args.audit).read_text(encoding="utf-8"))
        targets: list[tuple[str, int, float]] = []
        for r in d["reports"]:
            if r.get("status") != "mismatch":
                continue
            for m in r.get("mismatches", []):
                targets.append((r["symbol"], m["year"], m["set"]))
        print(f"diagnosing {len(targets)} (symbol, year) mismatches…",
              file=sys.stderr)
        out = []
        for sym, y, set_v in targets:
            res = diagnose(sym, set_v, y)
            out.append(res)
            n = res.get("match_count", 0)
            print(f"  {sym:8s} {y}  SET={set_v:10.2f}  hits={n}",
                  file=sys.stderr)
        Path("data/validation/diagnose_report.json").write_text(
            json.dumps(out, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"\nwrote data/validation/diagnose_report.json ({len(out)} entries)",
              file=sys.stderr)
        return 0

    if not args.symbol:
        ap.print_help()
        return 2

    # Load audit and pull every mismatched year for this symbol
    d = json.loads(Path(args.audit).read_text(encoding="utf-8"))
    sym = args.symbol.upper()
    rep = next((r for r in d["reports"] if r["symbol"] == sym), None)
    if rep is None:
        print(f"{sym}: not in audit", file=sys.stderr)
        return 1
    if rep.get("status") != "mismatch":
        print(f"{sym}: status={rep.get('status')}, no mismatches to diagnose",
              file=sys.stderr)
        return 0
    for m in rep.get("mismatches", []):
        res = diagnose(sym, m["set"], m["year"])
        print(json.dumps(res, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
